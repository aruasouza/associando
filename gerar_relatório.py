import pandas as pd
from openpyxl.chart import Reference,BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Border,Font,Side,PatternFill,Alignment
import os
from datetime import date

data_folder = 'dados'
output_folder = 'reports'

data_files = os.listdir(data_folder)

def create_table(df,writer,register,name,col):
    table = df.value_counts(name).rename('contagem').sort_values(ascending = False)
    table.to_excel(writer,sheet_name = tabels_sheet,index = True,startcol = col)
    register[name] = (len(table),col)

def create_chart(chart,datasheet,col,name,type_,style,width,height,table_len):
    data = Reference(datasheet,col + 2,2,col + 2,min(7,table_len + 1))
    cats = Reference(datasheet,col + 1,2,col + 1,min(7,table_len + 1))
    chart.type = type_
    chart.style = style
    chart.title = name
    chart.add_data(data, titles_from_data = False)
    chart.set_categories(cats)
    chart.legend = None
    chart.height = height
    chart.width = width
    chart.dataLabels = DataLabelList() 
    chart.dataLabels.showVal = True
    return chart

for file in data_files:
    tabels_sheet = 'tabelas'
    data_sheet = 'dados'
    socios_sheet = 'socios'
    in_path = os.path.join(data_folder,file)
    out_path = os.path.join(output_folder,file.replace('.csv','') + '_dashboard.xlsx')
    df = pd.read_csv(in_path,sep = ';')
    df['valor_capital_social'] = df['valor_capital_social'].apply(lambda x: float(x.replace(',','')) if not pd.isna(x) else None)
    with pd.ExcelWriter(out_path,engine = 'openpyxl') as writer:
        df.to_excel(writer,sheet_name = data_sheet,index = False)
        wb = writer.book
        tables = {}
        dash = wb.create_sheet("dashboard", 0)
        fill = PatternFill("solid", fgColor="00FFFFFF")
        for col in dash.iter_cols(min_row = 1,max_row = 60,min_col = 1,max_col = 30):
            for cell in col:
                cell.fill = fill
        title_font = Font(bold = True,size = 30)
        info_font = Font(bold = True,size = 12)
        cell = dash.cell(row = 1,column = 1,value = 'Dashboard')
        cell.font = title_font
        cell = dash.cell(row = 3,column = 1,value = 'UF:')
        cell.font = info_font
        cell = dash.cell(row = 4,column = 1,value = 'Cidade:')
        cell.font = info_font
        cell = dash.cell(row = 5,column = 1,value = 'Data:')
        cell.font = info_font
        cell = dash.cell(row = 3,column = 2,value = df['uf'].iloc[0])
        cell.font = info_font
        cell = dash.cell(row = 4,column = 2,value = df['nome_municipio'].iloc[0])
        cell.font = info_font
        cell = dash.cell(row = 5,column = 2,value = date.today().strftime('%d-%m-%Y'))
        cell.font = info_font

        row_contagem = 4
        row_capital = 5
        col_cards = 5
        cell = dash.cell(row = row_contagem,column = col_cards,value = 'Total de Empresas')
        cell.font = info_font
        cell = dash.cell(row = row_contagem,column = col_cards + 3,value = len(df))
        cell.font = info_font
        cell.alignment = Alignment('left')
        cell = dash.cell(row = row_capital,column = col_cards,value = 'Capital Social Médio')
        cell.font = info_font
        caps = df['valor_capital_social'].mean()
        cell = dash.cell(row = row_capital,column = col_cards + 3,value = f'R$ {caps:,.2f}')
        cell.font = info_font

        sheet = wb.create_sheet(tabels_sheet)
        name = 'descricao_cnae'
        col = 0
        create_table(df,writer,tables,name,col)
        template = BarChart()
        chart = create_chart(template,sheet,col,'Principais CNAES','bar',11,17,9,tables[name][0])
        template.y_axis
        dash.add_chart(chart, "A8")

        name = 'porte_empresa'
        col = 3
        create_table(df,writer,tables,name,col)
        template = BarChart()
        chart = create_chart(template,sheet,col,'Porte','col',12,8,9,tables[name][0])
        dash.add_chart(chart, "A27")

        name = 'bairro'
        col = 6
        create_table(df,writer,tables,name,col)
        template = BarChart()
        chart = create_chart(template,sheet,col,'Principais Bairros','col',11,17,9,tables[name][0])
        dash.add_chart(chart, "L8")

        name = 'natureza_juridica'
        col = 9
        create_table(df,writer,tables,name,col)
        template = BarChart()
        chart = create_chart(template,sheet,col,'Natureza Jurídica','col',12,14,9,tables[name][0])
        dash.add_chart(chart, "N27")

        name = 'quantidade_de_socios'
        col = 12
        create_table(df,writer,tables,name,col)
        template = BarChart()
        chart = create_chart(template,sheet,col,'Quantidade de Sócios','col',12,13,9,tables[name][0])
        dash.add_chart(chart, "F27")

        sheet = wb.create_sheet(socios_sheet)
        socio_name_cols = [col for col in df.columns if 'nome_socio' in col]
        socio_nascimento_cols = [col for col in df.columns if 'nascimento_socio' in col]
        socio_tel_cols = [col for col in df.columns if 'telefone_socio' in col]
        socio_representante_cols = [col for col in df.columns if 'nome_representante_legal_socio' in col]
        sheet.append(['cnpj_empresa','nome_socio','nascimento_socio','telefone_socio','nome_representante_legal_socio'])
        for i in df.index:
            for j in range(len(socio_name_cols)):
                if not pd.isna(df.loc[i,socio_name_cols[j]]):
                    row = df.loc[i]
                    sheet.append([row['cnpj'],row[socio_name_cols[j]],row[socio_nascimento_cols[j]],
                                  row[socio_tel_cols[j]],row[socio_representante_cols[j]]])
        
        font = Font(bold = False)
        side = Side(border_style = None)
        border = Border(left = side,right = side,top = side,bottom = side)
        sheet = wb[tabels_sheet]
        for table in tables:
            for i in range(2,tables[table][0] + 2):
                cell = sheet.cell(row = i,column = tables[table][1] + 1)
                cell.font = font
                cell.border = border
        for i in range(1,len(tables) * 3 + 1):
            cell = sheet.cell(row = 1,column = i)
            cell.font = font
            cell.border = border
        sheet = wb[data_sheet]
        for col in sheet.iter_cols(min_row=1, max_col=len(df.columns) + 1, max_row=1):
            for cell in col:
                cell.font = font
                cell.border = border